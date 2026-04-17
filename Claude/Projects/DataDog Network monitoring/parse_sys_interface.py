#!/usr/bin/env python3
"""
parse_sys_interface.py
-----------------------
Reads Fortigate `get system interface` CLI output and updates
FriendlyName / VPNDestination columns in SNMP_Interface_Inventory.xlsx.

HOW IT MAPS TUNNEL NAMES TO SNMP INDICES
-----------------------------------------
FortiOS lists interfaces in `get system interface` in creation order,
which matches SNMP ifIndex assignment order. IPsec VPN tunnels
(type=tunnel, non-zero IP) are extracted from the output in that order,
then matched positionally to the spreadsheet rows that use
match_field=index, sorted by their index number:

    CLI tunnel #1  →  spreadsheet row with lowest index
    CLI tunnel #2  →  spreadsheet row with 2nd-lowest index
    ... and so on.

USAGE
-----
    # Input from a saved text file:
    python parse_sys_interface.py <cli_output.txt> <device>

    # Input pasted directly (press Ctrl-D / Ctrl-Z when done):
    python parse_sys_interface.py - <device>

    <device>  can be a full sheet name ("DC - 10.1.1.1") or just the IP ("10.1.1.1")

OPTIONS
-------
    --xlsx PATH    Path to spreadsheet (default: SNMP_Interface_Inventory.xlsx in same folder)
    --dry-run      Print planned changes without saving
    --regen-yaml   After saving, run xlsx_to_yaml.py to regenerate conf.yaml

EXAMPLES
--------
    python parse_sys_interface.py dc_interfaces.txt 10.1.1.1
    python parse_sys_interface.py dc_interfaces.txt "DC - 10.1.1.1" --dry-run
    python parse_sys_interface.py - 10.1.1.1 --regen-yaml      # paste from stdin
"""

import sys
import re
import argparse
import subprocess
from pathlib import Path

import openpyxl

# ── Defaults ──────────────────────────────────────────────────────────────────
DEFAULT_XLSX = Path(__file__).parent / "SNMP_Interface_Inventory.xlsx"
DEFAULT_YAML = Path(__file__).parent / "conf.yaml"
XLSX_TO_YAML = Path(__file__).parent / "xlsx_to_yaml.py"

# System tunnel names that are never IPsec VPN tunnels
SYSTEM_TUNNELS = {"naf.root", "l2t.root", "ssl.root"}


# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_get_sys_interface(text):
    """
    Parse `get system interface` output into an ordered list of dicts.
    Each dict has: name, type, ip, status.
    The order matches SNMP ifIndex creation order.
    """
    interfaces = []
    current = None

    for line in text.splitlines():
        stripped = line.strip()

        # Skip firewall prompt lines (e.g. "FCTG-FG-DC-1 $")
        if stripped.endswith("$") and "get sys" not in stripped.lower():
            continue

        # New interface block: == [ name ]
        if stripped.startswith("== ["):
            end = stripped.find("]")
            if end != -1:
                name = stripped[4:end].strip()
                current = {"name": name, "type": None, "ip": "0.0.0.0", "status": None}
                interfaces.append(current)

        # Properties line — same block, starts with "name: <ifname>"
        # May be wrapped (long lines continue on the next line indented)
        elif current is not None:
            # Accumulate wrapped continuation lines into the properties
            props = stripped

            m = re.search(r"\btype:\s*(\S+)", props)
            if m:
                current["type"] = m.group(1)

            # Match "ip: A.B.C.D mask" — capture just the address
            m = re.search(r"\bip:\s*([\d.]+)\s+[\d.]+", props)
            if m:
                current["ip"] = m.group(1)

            m = re.search(r"\bstatus:\s*(\S+)", props)
            if m:
                current["status"] = m.group(1)

    return interfaces


def filter_vpn_tunnels(interfaces):
    """
    Return only IPsec VPN tunnel interfaces, in original order.

    Criteria:
      - type == 'tunnel'
      - name not in the known system tunnel list
      - ip is not 0.0.0.0 / None / empty  (system tunnels have no real IP)
    """
    return [
        iface for iface in interfaces
        if iface.get("type") == "tunnel"
        and iface["name"] not in SYSTEM_TUNNELS
        and iface.get("ip") not in ("0.0.0.0", None, "")
    ]


# ── Name helpers ──────────────────────────────────────────────────────────────

def derive_vpn_destination(tunnel_name):
    """
    Guess the VPNDestination tag value from the tunnel interface name.

    Strips trailing ordinal suffixes (_2, _2nd, _3rd …) then extracts
    the destination portion after the last _to_ or - separator.

    Examples:
        OPUS-FCTG        → FCTG
        OPUS-FCTG_2      → FCTG
        OPUS-Azure_2     → Azure
        OPUS_to_AIFP     → AIFP
        OPUS_to_AIFP_2   → AIFP
        OPUS_to_PFP_2nd  → PFP
        SMT_to_BIFP      → BIFP
        HQ_to_Azure_3rd  → Azure
    """
    # Strip trailing _2nd / _3rd / _2 / _3 etc.
    cleaned = re.sub(r"_\d+(st|nd|rd|th)$", "", tunnel_name, flags=re.IGNORECASE)
    cleaned = re.sub(r"_\d+$", "", cleaned)

    # SITE_to_DEST pattern
    m = re.search(r"_to_([A-Za-z0-9]+)$", cleaned, re.IGNORECASE)
    if m:
        return m.group(1)

    # SITE-DEST pattern  (e.g. OPUS-FCTG, OPUS-Azure)
    m = re.search(r"-([A-Za-z0-9]+)$", cleaned)
    if m:
        return m.group(1)

    return ""


# ── Spreadsheet helpers ────────────────────────────────────────────────────────

def find_sheet(workbook, device):
    """
    Locate the worksheet for <device>.
    <device> may be an exact sheet name or a substring (e.g. the IP).
    Raises ValueError if not found.
    """
    device = device.strip()

    if device in workbook.sheetnames:
        return workbook[device], device

    for name in workbook.sheetnames:
        if device in name:
            return workbook[name], name

    raise ValueError(
        f"No sheet found for '{device}'.\n"
        f"Available sheets: {', '.join(workbook.sheetnames)}"
    )


def read_headers(sheet):
    """
    Read row 2 as the column-header row.
    Returns dict: header_text → 1-based column index.
    """
    headers = {}
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=2, column=col).value
        if val:
            headers[str(val).strip()] = col
    return headers


def get_index_rows(sheet, headers):
    """
    Return all rows where match_field == 'index', sorted by index value (ascending).
    Each entry is (row_number, int_index).
    """
    mf_col = headers.get("Match Field")
    mv_col = headers.get("Match Value")
    if not mf_col or not mv_col:
        return []

    rows = []
    for row in range(3, sheet.max_row + 1):
        mf = sheet.cell(row=row, column=mf_col).value
        mv = sheet.cell(row=row, column=mv_col).value
        if str(mf).strip().lower() == "index" and mv is not None:
            try:
                rows.append((row, int(float(str(mv)))))
            except ValueError:
                pass

    return sorted(rows, key=lambda x: x[1])


def get_name_rows(sheet, headers):
    """
    Return a dict of interface_name → row_number for match_field == 'name' rows.
    """
    mf_col = headers.get("Match Field")
    mv_col = headers.get("Match Value")
    if not mf_col or not mv_col:
        return {}

    result = {}
    for row in range(3, sheet.max_row + 1):
        mf = sheet.cell(row=row, column=mf_col).value
        mv = sheet.cell(row=row, column=mv_col).value
        if str(mf).strip().lower() == "name" and mv:
            result[str(mv).strip()] = row
    return result


# ── Core update logic ─────────────────────────────────────────────────────────

def update_sheet(sheet, headers, all_interfaces, vpn_tunnels, dry_run=False):
    """
    Update FriendlyName and VPNDestination for index-based VPN rows.
    Also reports name-based entries whose interface is down or missing.

    MATCHING STRATEGY — two phases:
      Phase 1 (name match):  If a spreadsheet row already has a FriendlyName
                             that matches a CLI tunnel name, lock that pair in.
                             This is robust against extra tunnels in the CLI
                             output that aren't in the spreadsheet.
      Phase 2 (positional):  For rows still unmatched (FriendlyName is blank),
                             map them in order against the remaining unmatched
                             CLI tunnels (sorted index → CLI appearance order).

    Returns a list of human-readable change description strings.
    """
    fn_col = headers.get("Friendly Name")
    vd_col = headers.get("VPN Destination")
    mv_col = headers.get("Match Value")

    missing_cols = [c for c, v in [("Friendly Name", fn_col), ("VPN Destination", vd_col)]
                    if v is None]
    if missing_cols:
        print(f"  WARNING: column(s) not found in sheet: {', '.join(missing_cols)}")

    index_rows = get_index_rows(sheet, headers)   # [(row_num, int_index), ...] sorted by index
    name_rows  = get_name_rows(sheet, headers)

    # Build lookup: tunnel_name → tunnel dict
    tunnel_by_name = {t["name"]: t for t in vpn_tunnels}

    # ── Phase 1: name-based matching ─────────────────────────────────────────
    # For rows that already have a FriendlyName in the spreadsheet, match by name.
    matched_rows    = {}   # row_num → tunnel
    matched_tunnels = set()

    if fn_col:
        for row_num, idx in index_rows:
            existing_fn = str(sheet.cell(row=row_num, column=fn_col).value or "").strip()
            if existing_fn and existing_fn in tunnel_by_name:
                matched_rows[row_num] = tunnel_by_name[existing_fn]
                matched_tunnels.add(existing_fn)

    # ── Phase 2: positional matching for remaining unmatched rows ────────────
    unmatched_rows    = [(r, i) for r, i in index_rows if r not in matched_rows]
    unmatched_tunnels = [t for t in vpn_tunnels if t["name"] not in matched_tunnels]

    if unmatched_rows and unmatched_tunnels:
        if len(unmatched_tunnels) != len(unmatched_rows):
            print(
                f"\n  ⚠  Positional match: {len(unmatched_tunnels)} unmatched tunnel(s) "
                f"in CLI vs {len(unmatched_rows)} unmatched row(s) in spreadsheet.\n"
                f"     Pairing the first {min(len(unmatched_tunnels), len(unmatched_rows))} pairs."
            )
        for (row_num, idx), tunnel in zip(unmatched_rows, unmatched_tunnels):
            matched_rows[row_num] = tunnel

    # ── Apply matched pairs ───────────────────────────────────────────────────
    changes = []

    for row_num, idx in index_rows:
        if row_num not in matched_rows:
            fn = sheet.cell(row=row_num, column=fn_col).value if fn_col else ""
            changes.append(
                f"  Row {row_num}  index={idx:>3}  (no matching CLI tunnel — skipped, "
                f"current FriendlyName='{fn or ''}')"
            )
            continue

        tunnel = matched_rows[row_num]
        tname  = tunnel["name"]
        dest   = derive_vpn_destination(tname)
        row_changes = []

        if fn_col:
            existing = str(sheet.cell(row=row_num, column=fn_col).value or "").strip()
            if existing != tname:
                row_changes.append(f"FriendlyName '{existing}' → '{tname}'")
                if not dry_run:
                    sheet.cell(row=row_num, column=fn_col).value = tname

        if vd_col and dest:
            existing = str(sheet.cell(row=row_num, column=vd_col).value or "").strip()
            if existing != dest:
                row_changes.append(f"VPNDestination '{existing}' → '{dest}'")
                if not dry_run:
                    sheet.cell(row=row_num, column=vd_col).value = dest

        if row_changes:
            changes.append(
                f"  Row {row_num}  index={idx:>3}  tunnel='{tname}':  "
                + " | ".join(row_changes)
            )
        else:
            changes.append(
                f"  Row {row_num}  index={idx:>3}  tunnel='{tname}':  (already up to date)"
            )

    # ── Report CLI tunnels not matched to any spreadsheet row ─────────────────
    all_matched_tunnel_names = {t["name"] for t in matched_rows.values()}
    unmatched_cli = [t for t in vpn_tunnels if t["name"] not in all_matched_tunnel_names]
    if unmatched_cli:
        print(f"\n  ℹ  {len(unmatched_cli)} CLI tunnel(s) not mapped to any spreadsheet row:")
        for t in unmatched_cli:
            dest = derive_vpn_destination(t["name"])
            print(f"     → {t['name']:<40} VPNDestination would be: {dest or '(unknown)'}")
        print("     (Add rows to the spreadsheet if these should be monitored.)")

    # ── Check name-based interface entries against CLI output ─────────────────
    iface_by_name  = {i["name"]: i for i in all_interfaces}
    missing_ifaces = [n for n in name_rows if n not in iface_by_name]
    down_ifaces    = [n for n in name_rows
                      if n in iface_by_name and iface_by_name[n].get("status") == "down"]

    if missing_ifaces:
        print(f"\n  ⚠  Name-based interface(s) not found in CLI output:")
        for n in missing_ifaces:
            print(f"     - {n}")
    if down_ifaces:
        print(f"\n  ⚠  Name-based interface(s) currently DOWN:")
        for n in down_ifaces:
            print(f"     - {n}")

    return changes


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Update SNMP_Interface_Inventory.xlsx from 'get system interface' output.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "cli_output",
        help="Path to text file with CLI output, or '-' to read from stdin",
    )
    parser.add_argument(
        "device",
        help="Sheet name or IP of the device to update (e.g. '10.1.1.1' or 'DC - 10.1.1.1')",
    )
    parser.add_argument(
        "--xlsx",
        default=str(DEFAULT_XLSX),
        metavar="PATH",
        help=f"Path to the spreadsheet (default: {DEFAULT_XLSX.name} in same folder as this script)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned changes without saving",
    )
    parser.add_argument(
        "--regen-yaml",
        action="store_true",
        help="After saving, run xlsx_to_yaml.py to regenerate conf.yaml",
    )
    args = parser.parse_args()

    # ── Read CLI output ───────────────────────────────────────────────────────
    if args.cli_output == "-":
        print("Paste `get system interface` output below (Ctrl-D when done):\n")
        cli_text = sys.stdin.read()
    else:
        cli_path = Path(args.cli_output)
        if not cli_path.exists():
            print(f"ERROR: File not found: {cli_path}")
            sys.exit(1)
        cli_text = cli_path.read_text(encoding="utf-8", errors="replace")

    # ── Parse interfaces ──────────────────────────────────────────────────────
    all_interfaces = parse_get_sys_interface(cli_text)
    vpn_tunnels    = filter_vpn_tunnels(all_interfaces)

    print(f"Parsed {len(all_interfaces)} interfaces total.")
    print(f"Found {len(vpn_tunnels)} VPN tunnel(s) (type=tunnel, non-zero IP):\n")
    for i, t in enumerate(vpn_tunnels, 1):
        dest = derive_vpn_destination(t["name"])
        print(f"  {i:2d}. {t['name']:<35}  ip={t['ip']:<18} status={t['status'] or '?':<6}"
              f"  → VPNDest: {dest or '(unrecognised)'}")

    # ── Open spreadsheet ──────────────────────────────────────────────────────
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"\nERROR: Spreadsheet not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)

    try:
        sheet, sheet_name = find_sheet(wb, args.device)
    except ValueError as e:
        print(f"\nERROR: {e}")
        sys.exit(1)

    print(f"\nFound sheet: '{sheet_name}'")

    headers    = read_headers(sheet)
    index_rows = get_index_rows(sheet, headers)

    print(
        f"Index-based rows: {len(index_rows)}  "
        f"(indices: {[idx for _, idx in index_rows]})"
    )

    # ── Update ────────────────────────────────────────────────────────────────
    if args.dry_run:
        print("\n[DRY RUN — no changes will be saved]\n")

    changes = update_sheet(sheet, headers, all_interfaces, vpn_tunnels, dry_run=args.dry_run)

    actual_changes = [c for c in changes if "already up to date" not in c]
    up_to_date     = [c for c in changes if "already up to date" in c]

    if actual_changes:
        label = "Planned" if args.dry_run else "Applied"
        print(f"\n{label} {len(actual_changes)} change(s):")
        for c in actual_changes:
            print(c)
    else:
        print("\nNo changes needed — all fields already up to date.")

    if up_to_date:
        print(f"\n{len(up_to_date)} row(s) already correct (no change).")

    # ── Save ──────────────────────────────────────────────────────────────────
    if not args.dry_run:
        if actual_changes:
            wb.save(xlsx_path)
            print(f"\nSaved: {xlsx_path}")

            # Optionally regenerate conf.yaml
            if args.regen_yaml:
                if XLSX_TO_YAML.exists():
                    print(f"Regenerating conf.yaml via {XLSX_TO_YAML.name} ...")
                    result = subprocess.run(
                        [sys.executable, str(XLSX_TO_YAML), str(xlsx_path), str(DEFAULT_YAML)],
                        capture_output=True, text=True
                    )
                    if result.returncode == 0:
                        print(result.stdout.strip())
                        print(f"conf.yaml updated: {DEFAULT_YAML}")
                    else:
                        print(f"ERROR running xlsx_to_yaml.py:\n{result.stderr}")
                else:
                    print(f"WARNING: xlsx_to_yaml.py not found at {XLSX_TO_YAML}")
        else:
            print("\nNo changes to save.")
    else:
        print(f"\n(Run without --dry-run to apply changes.)")


if __name__ == "__main__":
    main()
