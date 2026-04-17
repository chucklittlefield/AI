#!/usr/bin/env python3
"""
analyze_fortigate_config.py
----------------------------
Parses a Fortigate full-configuration file ("show full-configuration" or
"sh full-configuration") and generates a multi-sheet Excel workbook for review.

Sheets produced:
  1. Summary          – firmware, model, hostname, object counts
  2. Interfaces       – all interfaces (IP, type, SNMP index, status, bandwidth)
  3. Zones            – zone names and member interfaces
  4. VPN Tunnels      – IPsec phase1 tunnels (WAN port, remote GW, proposal, DH)
  5. Static Routes    – destination, device, distance, priority, comment
  6. SD-WAN           – members, health-check probes, services
  7. Firewall Policies– ID, name, src/dst zone, src/dst addr, service, action, UTM
  8. Addresses        – address objects (subnet, FQDN, range, dynamic …)
  9. Address Groups   – group name → member list

Usage:
    python analyze_fortigate_config.py                         # uses defaults below
    python analyze_fortigate_config.py SIFPconfig.txt          # custom input
    python analyze_fortigate_config.py SIFPconfig.txt out.xlsx # custom input + output
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Defaults ──────────────────────────────────────────────────────────────────
DEFAULT_INPUT  = Path(__file__).parent / "SIFPconfig.txt"
DEFAULT_OUTPUT = Path(__file__).parent / "SIFP_Config_Analysis.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # dark blue  – column headers
C_SUBHEAD  = "2E75B6"   # mid blue   – sub-section headers
C_ALT1     = "DEEAF1"   # light blue – alternating row
C_ALT2     = "FFFFFF"   # white
C_WARN     = "FFE699"   # amber      – items needing attention
C_GOOD     = "E2EFDA"   # green      – enabled / up
C_BAD      = "FCE4D6"   # salmon     – disabled / down
C_ALL      = "FF9999"   # light red  – policy uses "all" in 1 of src/dst/service
C_ALL_MULTI = "C00000"  # dark red   – policy uses "all" in 2+ of src/dst/service
C_TITLE    = "1F4E79"   # same as header for sheet titles

FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_TITLE  = Font(bold=True, color="FFFFFF", size=12)
FONT_NORMAL = Font(size=10)
FONT_BOLD   = Font(bold=True, size=10)


# ══════════════════════════════════════════════════════════════════════════════
#  LOW-LEVEL PARSER
# ══════════════════════════════════════════════════════════════════════════════

def _strip_quotes(val: str) -> str:
    """Remove surrounding single or double quotes from a value."""
    v = val.strip()
    if (v.startswith('"') and v.endswith('"')) or \
       (v.startswith("'") and v.endswith("'")):
        return v[1:-1]
    return v


def _parse_set_line(line: str):
    """
    Parse a FortiOS 'set key value …' line.
    Returns (key, value_string) or None if not a set line.

    Multi-value fields (e.g. set srcaddr "A" "B" "C") are returned as
    comma-separated strings: "A, B, C".
    Single-value fields have their quotes stripped.
    """
    m = re.match(r'^\s+set\s+(\S+)\s+(.*)', line)
    if not m:
        return None
    key = m.group(1)
    raw = m.group(2).strip()

    # Extract all quoted tokens
    tokens = re.findall(r'"([^"]*)"', raw)
    if tokens:
        val = ", ".join(t for t in tokens)   # comma-joined, no quotes
    else:
        # Unquoted value — strip single quotes if present
        val = raw.strip("'")

    return key, val


def extract_top_section(text: str, section_name: str) -> str:
    """
    Extract the raw text of a 'config <section_name> … end' block.
    Works at any indentation level (top-level or nested inside another section).
    Returns empty string if not found.
    """
    pat = re.compile(
        r'^\s*config\s+' + re.escape(section_name) + r'\s*$',
        re.MULTILINE
    )
    m = pat.search(text)
    if not m:
        return ""

    start = m.end()
    depth = 1
    lines = text[start:].splitlines(keepends=True)
    chars = 0
    for line in lines:
        stripped = line.strip()
        if re.match(r'^config\b', stripped):
            depth += 1
        elif stripped == 'end':
            depth -= 1
            if depth == 0:
                return text[start: start + chars]
        chars += len(line)
    return text[start:]


def parse_entries(section_text: str) -> list:
    """
    Parse the edit/next blocks within a section.
    Returns a list of dicts, each with special key '__name__' for the
    edit identifier, plus all 'set key value' pairs.
    All values are raw strings.
    """
    entries = []
    current = None
    depth = 0           # nesting inside sub-configs
    lines = section_text.splitlines()

    for line in lines:
        stripped = line.strip()

        # Entering a sub-config block (inside an entry)
        if current is not None and re.match(r'^config\b', stripped) and depth == 0:
            depth += 1
            continue
        if depth > 0:
            if re.match(r'^config\b', stripped):
                depth += 1
            elif stripped == 'end':
                depth -= 1
            continue

        # New entry
        m = re.match(r'^edit\s+(.*)', stripped)
        if m:
            name = _strip_quotes(m.group(1))
            current = {'__name__': name}
            entries.append(current)
            continue

        # End of entry
        if stripped == 'next' and current is not None:
            current = None
            continue

        # Key-value pair inside entry
        if current is not None:
            parsed = _parse_set_line(line)
            if parsed:
                key, val = parsed
                current[key] = val

    return entries


def get_global_value(text: str, key: str) -> str:
    """Extract a single value from 'config system global'."""
    section = extract_top_section(text, "system global")
    for line in section.splitlines():
        parsed = _parse_set_line(line)
        if parsed and parsed[0] == key:
            return parsed[1]
    return ""


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION PARSERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_interfaces(text: str) -> list:
    section = extract_top_section(text, "system interface")
    entries = parse_entries(section)
    results = []
    for e in entries:
        ip_raw = e.get('ip', '0.0.0.0 0.0.0.0')
        parts  = ip_raw.split()
        ip     = parts[0] if parts else ''
        mask   = parts[1] if len(parts) > 1 else ''
        results.append({
            'Name':          e['__name__'],
            'Type':          e.get('type', ''),
            'IP':            ip,
            'Mask':          mask,
            'Status':        e.get('status', ''),
            'Role':          e.get('role', ''),
            'VLAN ID':       e.get('vlanid', ''),
            'Parent Iface':  e.get('interface', ''),
            'SNMP Index':    e.get('snmp-index', ''),
            'Description':   e.get('description', ''),
            'Alias':         e.get('alias', ''),
            'Allow Access':  e.get('allowaccess', ''),
            'VDOM':          e.get('vdom', ''),
            'In BW (bps)':   e.get('inbandwidth', '0'),
            'Out BW (bps)':  e.get('outbandwidth', '0'),
            'Est Up BW':     e.get('estimated-upstream-bandwidth', ''),
            'Est Down BW':   e.get('estimated-downstream-bandwidth', ''),
            'Speed':         e.get('speed', ''),
            'MTU Override':  e.get('mtu-override', ''),
            'Src Check':     e.get('src-check', ''),
        })
    return results


def parse_zones(text: str) -> list:
    section = extract_top_section(text, "system zone")
    entries = parse_entries(section)
    results = []
    for e in entries:
        results.append({
            'Zone Name':   e['__name__'],
            'Intrazone':   e.get('intrazone', ''),
            'Interfaces':  e.get('interface', ''),
            'Description': e.get('description', ''),
        })
    return results


def parse_vpn_phase1(text: str) -> list:
    section = extract_top_section(text, "vpn ipsec phase1-interface")
    entries = parse_entries(section)
    results = []
    for e in entries:
        results.append({
            'Tunnel Name':    e['__name__'],
            'WAN Interface':  e.get('interface', ''),
            'Remote GW':      e.get('remote-gw', ''),
            'Proposal':       e.get('proposal', ''),
            'DH Group':       e.get('dhgrp', ''),
            'IKE Version':    e.get('ike-version', '1'),
            'Auth Method':    e.get('authmethod', ''),
            'Mode':           e.get('mode', ''),
            'DPD':            e.get('dpd', ''),
            'NAT Traversal':  e.get('nattraversal', ''),
            'Type':           e.get('type', ''),
            'Comments':       e.get('comments', ''),
        })
    return results


def parse_vpn_phase2(text: str) -> list:
    section = extract_top_section(text, "vpn ipsec phase2-interface")
    entries = parse_entries(section)
    results = []
    for e in entries:
        results.append({
            'Phase2 Name':   e['__name__'],
            'Phase1 Name':   e.get('phase1name', ''),
            'Src Subnet':    e.get('src-subnet', ''),
            'Dst Subnet':    e.get('dst-subnet', ''),
            'Proposal':      e.get('proposal', ''),
            'PFS':           e.get('pfs', ''),
            'Auto Negotiate':e.get('auto-negotiate', ''),
            'Keylife (s)':   e.get('keylifeseconds', ''),
        })
    return results


def parse_services(text: str) -> list:
    section = extract_top_section(text, "firewall service custom")
    entries = parse_entries(section)
    results = []
    for e in entries:
        # Build a compact port summary
        tcp = e.get('tcp-portrange', '')
        udp = e.get('udp-portrange', '')
        sctp = e.get('sctp-portrange', '')
        ports_parts = []
        if tcp:  ports_parts.append(f"TCP:{tcp}")
        if udp:  ports_parts.append(f"UDP:{udp}")
        if sctp: ports_parts.append(f"SCTP:{sctp}")
        results.append({
            'Name':       e['__name__'],
            'Category':   e.get('category', ''),
            'Protocol':   e.get('protocol', ''),
            'Ports':      '  '.join(ports_parts),
            'TCP Ports':  tcp,
            'UDP Ports':  udp,
            'Helper':     e.get('helper', ''),
            'FQDN':       e.get('fqdn', ''),
            'IP Range':   e.get('iprange', ''),
            'Proxy':      e.get('proxy', ''),
            'Comment':    e.get('comment', ''),
        })
    return results


def parse_service_groups(text: str) -> list:
    section = extract_top_section(text, "firewall service group")
    entries = parse_entries(section)
    results = []
    for e in entries:
        results.append({
            'Group Name': e['__name__'],
            'Members':    e.get('member', ''),
            'Proxy':      e.get('proxy', ''),
            'Comment':    e.get('comment', ''),
        })
    return results


def parse_static_routes(text: str) -> list:
    section = extract_top_section(text, "router static")
    entries = parse_entries(section)
    results = []
    for e in entries:
        dst_raw = e.get('dst', '0.0.0.0 0.0.0.0')
        parts   = dst_raw.split()
        dst     = parts[0] if parts else ''
        mask    = parts[1] if len(parts) > 1 else ''
        results.append({
            'ID':        e['__name__'],
            'Status':    e.get('status', 'enable'),
            'Dst Network': dst,
            'Dst Mask':  mask,
            'Gateway':   e.get('gateway', ''),
            'Device':    e.get('device', ''),
            'SD-WAN Zone': e.get('sdwan-zone', ''),
            'Distance':  e.get('distance', ''),
            'Priority':  e.get('priority', ''),
            'Weight':    e.get('weight', ''),
            'Blackhole': e.get('blackhole', ''),
            'Comment':   e.get('comment', ''),
        })
    return sorted(results, key=lambda r: (r['Distance'] or '999', r['Priority'] or '999'))


def parse_sdwan(text: str) -> dict:
    """Returns dict with keys 'members', 'health_checks', 'services'."""
    section = extract_top_section(text, "system sdwan")

    # ── Members ──────────────────────────────────────────────────────────────
    members_txt = extract_top_section(section, "members")
    members = []
    for e in parse_entries(members_txt):
        members.append({
            'ID':         e['__name__'],
            'Interface':  e.get('interface', ''),
            'Zone':       e.get('zone', ''),
            'Gateway':    e.get('gateway', ''),
            'Priority':   e.get('priority', ''),
            'Cost':       e.get('cost', ''),
            'Status':     e.get('status', ''),
            'Comment':    e.get('comment', ''),
        })

    # ── Health Checks ─────────────────────────────────────────────────────────
    hc_txt = extract_top_section(section, "health-check")
    health_checks = []
    for e in parse_entries(hc_txt):
        health_checks.append({
            'Name':          e['__name__'],
            'Server':        e.get('server', ''),
            'Protocol':      e.get('protocol', 'ping'),
            'Interval (ms)': e.get('interval', ''),
            'Timeout (ms)':  e.get('probe-timeout', ''),
            'Fail Time':     e.get('failtime', ''),
            'Recover Time':  e.get('recoverytime', ''),
            'Probe Count':   e.get('probe-count', ''),
            'System DNS':    e.get('system-dns', ''),
            'Update Static Route': e.get('update-static-route', ''),
        })

    # ── Services ──────────────────────────────────────────────────────────────
    svc_txt = extract_top_section(section, "service")
    services = []
    for e in parse_entries(svc_txt):
        services.append({
            'ID':              e['__name__'],
            'Name':            e.get('name', ''),
            'Mode':            e.get('mode', ''),
            'Load Balance Mode': e.get('load-balance-mode', ''),
            'Priority Members':e.get('priority-members', ''),
            'Health Check':    e.get('health-check', ''),
            'Min Quality Members': e.get('minimum-sla-meet-members', ''),
            'DST':             e.get('dst', ''),
            'Src':             e.get('src', ''),
            'Internet Service':e.get('internet-service-name', ''),
        })

    return {'members': members, 'health_checks': health_checks, 'services': services}


def parse_policies(text: str) -> list:
    section = extract_top_section(text, "firewall policy")
    entries = parse_entries(section)
    results = []
    for e in entries:
        results.append({
            'ID':           e['__name__'],
            'Name':         e.get('name', ''),
            'Status':       e.get('status', 'enable'),
            'Src Interface':e.get('srcintf', ''),
            'Dst Interface':e.get('dstintf', ''),
            'Src Address':  e.get('srcaddr', ''),
            'Dst Address':  e.get('dstaddr', ''),
            'Service':      e.get('service', ''),
            'Action':       e.get('action', ''),
            'NAT':          e.get('nat', ''),
            'NAT IP':       e.get('natip', ''),
            'Schedule':     e.get('schedule', ''),
            'UTM Status':   e.get('utm-status', ''),
            'AV Profile':   e.get('av-profile', ''),
            'IPS Sensor':   e.get('ips-sensor', ''),
            'Web Filter':   e.get('webfilter-profile', ''),
            'App Control':  e.get('application-list', ''),
            'SSL Inspect':  e.get('ssl-ssh-profile', ''),
            'Log Traffic':  e.get('logtraffic', ''),
            'Comments':     e.get('comments', ''),
        })
    # Sort by policy ID numerically
    results.sort(key=lambda r: int(r['ID']) if str(r['ID']).isdigit() else 0)
    return results


def parse_addresses(text: str) -> list:
    section = extract_top_section(text, "firewall address")
    entries = parse_entries(section)
    results = []
    for e in entries:
        addr_type = e.get('type', 'ipmask')
        if addr_type == 'ipmask':
            value = e.get('subnet', '')
        elif addr_type == 'fqdn':
            value = e.get('fqdn', '')
        elif addr_type == 'iprange':
            value = f"{e.get('start-ip','')} – {e.get('end-ip','')}"
        elif addr_type == 'dynamic':
            value = f"dynamic ({e.get('sub-type','')})"
        elif addr_type == 'geography':
            value = e.get('country', '')
        else:
            value = ''
        results.append({
            'Name':        e['__name__'],
            'Type':        addr_type,
            'Value':       value,
            'Interface':   e.get('associated-interface', ''),
            'Comment':     e.get('comment', ''),
            'Fabric Obj':  e.get('fabric-object', ''),
        })
    return results


def parse_addrgrps(text: str) -> list:
    section = extract_top_section(text, "firewall addrgrp")
    entries = parse_entries(section)
    results = []
    for e in entries:
        results.append({
            'Group Name': e['__name__'],
            'Members':    e.get('member', ''),
            'Comment':    e.get('comment', ''),
        })
    return results


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _thin_border() -> Border:
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _write_title(ws, title: str, ncols: int):
    """Write a full-width coloured title row."""
    ws.append([title])
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.fill  = _fill(C_TITLE)
    cell.font  = FONT_TITLE
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 20

def _write_headers(ws, headers: list):
    """Write a bold blue header row."""
    ws.append(headers)
    row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = _fill(C_HEADER)
        cell.font      = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[row].height = 18

def _write_row(ws, values: list, alt: bool = False, highlight: str = None):
    """Write a data row with optional alternating colour or highlight."""
    ws.append(values)
    row = ws.max_row
    color = highlight if highlight else (C_ALT1 if alt else C_ALT2)
    # Use white font on dark backgrounds for readability
    font = Font(bold=True, color="FFFFFF", size=10) if highlight == C_ALL_MULTI else FONT_NORMAL
    for col in range(1, len(values) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = _fill(color)
        cell.font      = font
        cell.alignment = Alignment(vertical='center', wrap_text=False)
        cell.border    = _thin_border()

def _auto_width(ws, min_w=8, max_w=45):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def _freeze(ws, row=3, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_password_policy(text: str) -> dict:
    """Parse config system password-policy settings."""
    defaults = {
        'Status':          'disable',
        'Min Length':      '8',
        'Min Upper':       '0',
        'Min Lower':       '0',
        'Min Number':      '0',
        'Min Non-Alpha':   '0',
        'Expire Status':   'disable',
        'Expire Day':      'N/A',
        'Reuse Password':  'enable',
        'Change 4 chars':  'disable',
    }
    m = re.search(r'config system password-policy\n(.*?)\nend', text, re.DOTALL)
    if not m:
        return defaults
    body = m.group(1)
    def g(key):
        hit = re.search(rf'set {key}\s+(\S+)', body)
        return hit.group(1) if hit else None
    result = dict(defaults)
    for attr, key in [
        ('Status',         'status'),
        ('Min Length',     'minimum-length'),
        ('Min Upper',      'min-upper-case-letter'),
        ('Min Lower',      'min-lower-case-letter'),
        ('Min Number',     'min-number'),
        ('Min Non-Alpha',  'min-non-alphanumeric'),
        ('Expire Status',  'expire-status'),
        ('Expire Day',     'expire-day'),
        ('Reuse Password', 'reuse-password'),
        ('Change 4 chars', 'change-4-characters'),
    ]:
        v = g(key)
        if v is not None:
            result[attr] = v
    return result


def parse_admin_users(text: str) -> list:
    """Parse config system admin block, returning one dict per user."""
    users = []
    lines = text.splitlines()
    start = next((i for i, l in enumerate(lines) if l.strip() == 'config system admin'), None)
    if start is None:
        return users
    # Find the matching end at depth 1
    end = None
    depth = 1
    for i in range(start + 1, len(lines)):
        s = lines[i].strip()
        if re.match(r'^config\b', s):
            depth += 1
        elif s == 'end':
            depth -= 1
            if depth == 0:
                end = i
                break
    if end is None:
        return users

    block = '\n'.join(lines[start + 1:end])
    entries = re.split(r'\n    next\b', block)

    for entry in entries:
        nm = re.search(r'edit "([^"]+)"', entry)
        if not nm:
            continue
        def g(key, default=''):
            hit = re.search(rf'set {key}\s+"?([^"\n]+)"?', entry)
            return hit.group(1).strip() if hit else default
        def gq(key, default=''):
            hit = re.search(rf'set {key}\s+"([^"]+)"', entry)
            return hit.group(1).strip() if hit else default

        # Collect trust hosts (skip 0.0.0.0 255.255.255.255 = "any")
        trusts = []
        for n in range(1, 11):
            th = re.search(rf'set trusthost{n}\s+([\d./]+ [\d./]+)', entry)
            if th:
                ip, mask = th.group(1).split()
                if ip != '0.0.0.0':
                    trusts.append(f"{ip}/{mask}")
        trust_str = ', '.join(trusts) if trusts else 'any'

        users.append({
            'Username':    nm.group(1),
            'Profile':     gq('accprofile', 'N/A'),
            'VDOM':        gq('vdom', 'root'),
            '2FA':         g('two-factor', 'none'),
            'Remote Auth': g('remote-auth', 'disable'),
            'Wildcard':    g('wildcard', 'disable'),
            'Admin Timeout': g('gui-ignore-release-overview-version', ''),
            'Allowed Hosts': trust_str,
            'Email':       gq('email-to', ''),
        })
    return users


def build_summary(wb, text: str, counts: dict, src_file: str, pwd_policy: dict):
    ws = wb.create_sheet("Summary")

    hostname = get_global_value(text, "hostname")
    alias    = get_global_value(text, "alias")
    tz       = get_global_value(text, "timezone")
    admintimeout = get_global_value(text, "admintimeout")

    # Extract firmware from first comment line
    fw_match = re.search(r'#config-version=(\S+)', text)
    firmware = fw_match.group(1) if fw_match else "unknown"
    fw_parts = firmware.split(':')[0]

    _write_title(ws, "Fortigate Configuration Analysis", 3)
    ws.append([])

    info = [
        ("Source File",       src_file),
        ("Hostname",          hostname),
        ("Device Alias",      alias),
        ("Firmware",          fw_parts),
        ("Timezone",          tz),
        ("Admin Session Timeout", f"{admintimeout} min" if admintimeout else "N/A"),
    ]
    for label, val in info:
        ws.append([label, val])
        r = ws.max_row
        ws.cell(r, 1).font   = FONT_BOLD
        ws.cell(r, 1).fill   = _fill(C_ALT1)
        ws.cell(r, 2).font   = FONT_NORMAL

    ws.append([])
    ws.append(["Object", "Count"])
    r = ws.max_row
    ws.cell(r, 1).fill = _fill(C_HEADER); ws.cell(r, 1).font = FONT_HEADER
    ws.cell(r, 2).fill = _fill(C_HEADER); ws.cell(r, 2).font = FONT_HEADER

    for label, count in counts.items():
        ws.append([label, count])
        r = ws.max_row
        ws.cell(r, 1).font = FONT_NORMAL
        ws.cell(r, 2).font = FONT_BOLD

    # ── Password Policy section ───────────────────────────────────────────────
    ws.append([])
    ws.append(["Password Policy Settings", ""])
    r = ws.max_row
    ws.cell(r, 1).fill = _fill(C_SUBHEAD); ws.cell(r, 1).font = FONT_HEADER
    ws.cell(r, 2).fill = _fill(C_SUBHEAD); ws.cell(r, 2).font = FONT_HEADER

    policy_display = [
        ("Policy Status",          pwd_policy['Status']),
        ("Minimum Length",         pwd_policy['Min Length']),
        ("Min Uppercase Letters",  pwd_policy['Min Upper']),
        ("Min Lowercase Letters",  pwd_policy['Min Lower']),
        ("Min Numeric Characters", pwd_policy['Min Number']),
        ("Min Non-Alphanumeric",   pwd_policy['Min Non-Alpha']),
        ("Password Expiry",        pwd_policy['Expire Status']),
        ("Expiry Period (days)",   pwd_policy['Expire Day']),
        ("Allow Password Reuse",   pwd_policy['Reuse Password']),
        ("Must Change 4 Chars",    pwd_policy['Change 4 chars']),
    ]
    for i, (label, val) in enumerate(policy_display):
        # Flag concerning settings in amber
        concern = (
            (label == "Policy Status" and val == 'disable') or
            (label == "Minimum Length" and int(val) < 12) or
            (label == "Allow Password Reuse" and val == 'enable') or
            (label == "Password Expiry" and val == 'disable')
        )
        ws.append([label, val])
        r = ws.max_row
        fill_color = C_WARN if concern else (C_ALT1 if i % 2 == 0 else C_ALT2)
        ws.cell(r, 1).font  = FONT_BOLD
        ws.cell(r, 1).fill  = _fill(fill_color)
        ws.cell(r, 2).font  = FONT_NORMAL
        ws.cell(r, 2).fill  = _fill(fill_color)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20


def build_admin_users(wb, users: list):
    ws = wb.create_sheet("Admin Users")
    headers = ['Username', 'Access Profile', 'VDOM', '2FA', 'Remote Auth',
               'Wildcard', 'Allowed Hosts', 'Email']
    _write_title(ws, f"Admin Users  ({len(users)} total)", len(headers))
    _write_headers(ws, headers)

    # Profile → friendly label map
    profile_labels = {
        'super_admin':          'Super Admin (Full Access)',
        'super_admin_readonly': 'Read Only (Super Admin)',
        'prof_admin':           'Profile Admin (Full Access)',
        'Read_Only':            'Read Only',
    }

    for i, row in enumerate(users):
        profile = row['Profile']
        friendly_profile = profile_labels.get(profile, profile)
        # Highlight: read-only = green, super_admin = amber, unknown = white
        if 'readonly' in profile.lower() or profile == 'Read_Only':
            highlight = C_GOOD
        elif profile in ('super_admin', 'prof_admin'):
            highlight = C_WARN
        else:
            highlight = None

        _write_row(ws, [
            row['Username'],
            friendly_profile,
            row['VDOM'],
            row['2FA'],
            row['Remote Auth'],
            row['Wildcard'],
            row['Allowed Hosts'],
            row['Email'],
        ], alt=i % 2 == 0, highlight=highlight)

    _auto_width(ws)
    _freeze(ws)


def build_interfaces(wb, interfaces: list):
    ws = wb.create_sheet("Interfaces")
    headers = [
        'Name','Type','IP','Mask','Status','Role','VLAN ID','Parent Iface',
        'SNMP Index','Allow Access','In BW (bps)','Out BW (bps)',
        'Est Up BW','Est Down BW','Speed','Description','Alias','VDOM'
    ]
    _write_title(ws, f"Interfaces  ({len(interfaces)} total)", len(headers))
    _write_headers(ws, headers)

    # Sort: physical first, then by type, then name
    type_order = {'physical':0,'aggregate':1,'hard-switch':2,'vlan':3,'tunnel':4,'loopback':5}
    interfaces.sort(key=lambda r: (type_order.get(r['Type'], 9), r['Name']))

    for i, row in enumerate(interfaces):
        status = row['Status']
        highlight = None
        if status == 'down':
            highlight = C_BAD
        elif status == 'up':
            highlight = C_GOOD if row['Type'] in ('physical','aggregate') else None

        _write_row(ws, [
            row['Name'], row['Type'], row['IP'], row['Mask'],
            row['Status'], row['Role'], row['VLAN ID'], row['Parent Iface'],
            row['SNMP Index'], row['Allow Access'],
            row['In BW (bps)'], row['Out BW (bps)'],
            row['Est Up BW'], row['Est Down BW'],
            row['Speed'], row['Description'], row['Alias'], row['VDOM']
        ], alt=i % 2 == 0, highlight=highlight)

    _auto_width(ws)
    _freeze(ws)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_zones(wb, zones: list):
    ws = wb.create_sheet("Zones")
    headers = ['Zone Name', 'Intrazone', 'Member Interfaces', 'Description']
    _write_title(ws, f"Zones  ({len(zones)} total)", len(headers))
    _write_headers(ws, headers)

    for i, row in enumerate(zones):
        _write_row(ws, [
            row['Zone Name'], row['Intrazone'],
            row['Interfaces'], row['Description']
        ], alt=i % 2 == 0)

    _auto_width(ws)
    _freeze(ws)


def build_vpn(wb, phase1: list, phase2: list):
    ws = wb.create_sheet("VPN Tunnels")

    # ── Phase 1 ───────────────────────────────────────────────────────────────
    h1 = ['Tunnel Name','WAN Interface','Remote GW','Proposal','DH Group',
          'IKE Ver','Auth','Mode','DPD','NAT-T','Type','Comments']
    _write_title(ws, f"IPsec Phase 1 Tunnels  ({len(phase1)} tunnels)", len(h1))
    _write_headers(ws, h1)
    for i, row in enumerate(phase1):
        _write_row(ws, [
            row['Tunnel Name'], row['WAN Interface'], row['Remote GW'],
            row['Proposal'], row['DH Group'], row['IKE Version'],
            row['Auth Method'], row['Mode'], row['DPD'],
            row['NAT Traversal'], row['Type'], row['Comments']
        ], alt=i % 2 == 0)

    ws.append([])

    # ── Phase 2 ───────────────────────────────────────────────────────────────
    h2 = ['Phase2 Name','Phase1 Name','Src Subnet','Dst Subnet',
          'Proposal','PFS','Auto Negotiate','Keylife (s)']
    ws.append([f"IPsec Phase 2  ({len(phase2)} entries)"])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(h2))
    ws.cell(r, 1).fill = _fill(C_SUBHEAD)
    ws.cell(r, 1).font = Font(bold=True, color="FFFFFF", size=11)
    _write_headers(ws, h2)
    for i, row in enumerate(phase2):
        _write_row(ws, [
            row['Phase2 Name'], row['Phase1 Name'],
            row['Src Subnet'], row['Dst Subnet'],
            row['Proposal'], row['PFS'],
            row['Auto Negotiate'], row['Keylife (s)']
        ], alt=i % 2 == 0)

    _auto_width(ws)
    _freeze(ws)


def build_static_routes(wb, routes: list):
    ws = wb.create_sheet("Static Routes")
    headers = ['ID','Status','Dst Network','Dst Mask','Gateway','Device',
               'SD-WAN Zone','Distance','Priority','Weight','Blackhole','Comment']
    _write_title(ws, f"Static Routes  ({len(routes)} total)", len(headers))
    _write_headers(ws, headers)

    for i, row in enumerate(routes):
        highlight = C_BAD if row['Status'] == 'disable' else None
        _write_row(ws, [
            row['ID'], row['Status'], row['Dst Network'], row['Dst Mask'],
            row['Gateway'], row['Device'], row['SD-WAN Zone'],
            row['Distance'], row['Priority'], row['Weight'],
            row['Blackhole'], row['Comment']
        ], alt=i % 2 == 0, highlight=highlight)

    _auto_width(ws)
    _freeze(ws)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_sdwan(wb, sdwan: dict):
    ws = wb.create_sheet("SD-WAN")

    # ── Members ───────────────────────────────────────────────────────────────
    hm = ['ID','Interface','Zone','Gateway','Priority','Cost','Status','Comment']
    members = sdwan['members']
    _write_title(ws, f"SD-WAN Members  ({len(members)})", len(hm))
    _write_headers(ws, hm)
    for i, row in enumerate(members):
        highlight = C_BAD if row['Status'] == 'disable' else None
        _write_row(ws, [
            row['ID'], row['Interface'], row['Zone'], row['Gateway'],
            row['Priority'], row['Cost'], row['Status'], row['Comment']
        ], alt=i % 2 == 0, highlight=highlight)

    ws.append([])

    # ── Health Checks ─────────────────────────────────────────────────────────
    hh = ['Name','Server','Protocol','Interval (ms)','Timeout (ms)',
          'Fail Time','Recover Time','Probe Count','System DNS','Update Static Route']
    hcs = sdwan['health_checks']
    ws.append([f"Health Checks  ({len(hcs)})"])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(hh))
    ws.cell(r, 1).fill = _fill(C_SUBHEAD); ws.cell(r, 1).font = Font(bold=True, color="FFFFFF", size=11)
    _write_headers(ws, hh)
    for i, row in enumerate(hcs):
        _write_row(ws, [
            row['Name'], row['Server'], row['Protocol'],
            row['Interval (ms)'], row['Timeout (ms)'],
            row['Fail Time'], row['Recover Time'], row['Probe Count'],
            row['System DNS'], row['Update Static Route']
        ], alt=i % 2 == 0)

    ws.append([])

    # ── Services ──────────────────────────────────────────────────────────────
    hs = ['ID','Name','Mode','Load Balance Mode','Priority Members',
          'Health Check','DST','Src','Internet Service']
    svcs = sdwan['services']
    ws.append([f"SD-WAN Services / Rules  ({len(svcs)})"])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(hs))
    ws.cell(r, 1).fill = _fill(C_SUBHEAD); ws.cell(r, 1).font = Font(bold=True, color="FFFFFF", size=11)
    _write_headers(ws, hs)
    for i, row in enumerate(svcs):
        _write_row(ws, [
            row['ID'], row['Name'], row['Mode'], row['Load Balance Mode'],
            row['Priority Members'], row['Health Check'],
            row['DST'], row['Src'], row['Internet Service']
        ], alt=i % 2 == 0)

    _auto_width(ws)
    _freeze(ws)


def _has_all(value: str) -> bool:
    """Return True if 'all' appears as one of the comma-separated values (case-insensitive)."""
    if not value:
        return False
    return any(p.strip().lower() == 'all' for p in str(value).split(','))


def build_policies(wb, policies: list):
    ws = wb.create_sheet("Firewall Policies")
    headers = [
        'ID','Name','Status','Src Interface','Dst Interface',
        'Src Address','Dst Address','Service','Action','NAT',
        'UTM','AV Profile','IPS Sensor','Web Filter','App Control',
        'SSL Inspect','Log Traffic','Schedule','Comments'
    ]

    single_all_count = sum(
        1 for p in policies
        if sum([_has_all(p['Src Address']), _has_all(p['Dst Address']), _has_all(p['Service'])]) == 1
    )
    multi_all_count = sum(
        1 for p in policies
        if sum([_has_all(p['Src Address']), _has_all(p['Dst Address']), _has_all(p['Service'])]) >= 2
    )
    _write_title(ws,
        f"Firewall Policies  ({len(policies)} total  |  "
        f"{single_all_count} with 'all' in 1 col — light red  |  "
        f"{multi_all_count} with 'all' in 2+ cols — dark red)",
        len(headers))
    _write_headers(ws, headers)

    for i, row in enumerate(policies):
        action = row['Action']
        status = row['Status']
        all_field_count = sum([
            _has_all(row['Src Address']),
            _has_all(row['Dst Address']),
            _has_all(row['Service'])
        ])

        # Priority: dark red (2+ "all" cols) > light red (1 "all" col) > disabled salmon > deny amber
        if all_field_count >= 2:
            highlight = C_ALL_MULTI
        elif all_field_count == 1:
            highlight = C_ALL
        elif status == 'disable':
            highlight = C_BAD
        elif action == 'deny':
            highlight = C_WARN
        else:
            highlight = None

        _write_row(ws, [
            row['ID'], row['Name'], row['Status'],
            row['Src Interface'], row['Dst Interface'],
            row['Src Address'], row['Dst Address'],
            row['Service'], row['Action'], row['NAT'],
            row['UTM Status'], row['AV Profile'], row['IPS Sensor'],
            row['Web Filter'], row['App Control'], row['SSL Inspect'],
            row['Log Traffic'], row['Schedule'], row['Comments']
        ], alt=i % 2 == 0, highlight=highlight)

    _auto_width(ws)
    _freeze(ws)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_addresses(wb, addresses: list):
    ws = wb.create_sheet("Addresses")
    headers = ['Name','Type','Value / Subnet / FQDN','Interface','Comment','Fabric Object']
    _write_title(ws, f"Firewall Address Objects  ({len(addresses)} total)", len(headers))
    _write_headers(ws, headers)

    type_color = {
        'ipmask': None, 'fqdn': C_ALT1, 'iprange': C_WARN,
        'dynamic': "E2CFED", 'geography': "D9E1F2"
    }
    for i, row in enumerate(addresses):
        highlight = type_color.get(row['Type'])
        _write_row(ws, [
            row['Name'], row['Type'], row['Value'],
            row['Interface'], row['Comment'], row['Fabric Obj']
        ], alt=i % 2 == 0, highlight=highlight)

    _auto_width(ws)
    _freeze(ws)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_services(wb, services: list, svc_groups: list):
    ws = wb.create_sheet("Services")

    # ── Custom service objects ────────────────────────────────────────────────
    # Separate built-in (no custom ports / protocol=IP) from custom
    custom   = [s for s in services if s['TCP Ports'] or s['UDP Ports'] or s['FQDN']]
    builtin  = [s for s in services if s not in custom]

    h1 = ['Name', 'Category', 'Protocol', 'TCP Ports', 'UDP Ports',
          'Helper', 'FQDN', 'IP Range', 'Comment']
    _write_title(ws, f"Custom Service Objects  ({len(services)} total  |  "
                     f"{len(custom)} with explicit ports  |  {len(builtin)} built-in/protocol-based)",
                 len(h1))
    _write_headers(ws, h1)

    # Custom port-defined services first, then built-ins
    for i, row in enumerate(custom + builtin):
        highlight = None if row in custom else C_ALT1
        _write_row(ws, [
            row['Name'], row['Category'], row['Protocol'],
            row['TCP Ports'], row['UDP Ports'],
            row['Helper'], row['FQDN'], row['IP Range'], row['Comment']
        ], alt=i % 2 == 0, highlight=highlight)

    ws.append([])

    # ── Service groups ────────────────────────────────────────────────────────
    h2 = ['Group Name', 'Members', 'Proxy', 'Comment']
    ws.append([f"Service Groups  ({len(svc_groups)})"])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(h2))
    ws.cell(r, 1).fill = _fill(C_SUBHEAD)
    ws.cell(r, 1).font = Font(bold=True, color="FFFFFF", size=11)

    _write_headers(ws, h2)
    for i, row in enumerate(svc_groups):
        _write_row(ws, [
            row['Group Name'], row['Members'], row['Proxy'], row['Comment']
        ], alt=i % 2 == 0)

    _auto_width(ws)
    ws.column_dimensions['B'].width = 55   # Members column can be wide
    _freeze(ws)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(h1))}{len(services) + 2}"


def build_addrgrps(wb, groups: list):
    ws = wb.create_sheet("Address Groups")
    headers = ['Group Name','Members','Comment']
    _write_title(ws, f"Address Groups  ({len(groups)} total)", len(headers))
    _write_headers(ws, headers)

    for i, row in enumerate(groups):
        _write_row(ws, [
            row['Group Name'], row['Members'], row['Comment']
        ], alt=i % 2 == 0)

    _auto_width(ws)
    ws.column_dimensions['B'].width = 60
    _freeze(ws)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # ── Argument handling ──────────────────────────────────────────────────────
    args = sys.argv[1:]
    input_path  = Path(args[0]) if len(args) >= 1 else DEFAULT_INPUT
    output_path = Path(args[1]) if len(args) >= 2 else DEFAULT_OUTPUT

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    print(f"Reading: {input_path}")
    text = input_path.read_text(encoding="utf-8", errors="replace")

    # ── Parse all sections ────────────────────────────────────────────────────
    print("Parsing interfaces …")
    interfaces = parse_interfaces(text)

    print("Parsing zones …")
    zones = parse_zones(text)

    print("Parsing VPN tunnels …")
    phase1 = parse_vpn_phase1(text)
    phase2 = parse_vpn_phase2(text)

    print("Parsing static routes …")
    routes = parse_static_routes(text)

    print("Parsing SD-WAN …")
    sdwan = parse_sdwan(text)

    print("Parsing firewall policies …")
    policies = parse_policies(text)

    print("Parsing services …")
    services   = parse_services(text)
    svc_groups = parse_service_groups(text)

    print("Parsing address objects …")
    addresses = parse_addresses(text)

    print("Parsing address groups …")
    addrgrps = parse_addrgrps(text)

    print("Parsing admin users …")
    admin_users = parse_admin_users(text)

    print("Parsing password policy …")
    pwd_policy = parse_password_policy(text)

    counts = {
        "Interfaces":         len(interfaces),
        "  – Physical":       sum(1 for i in interfaces if i['Type'] == 'physical'),
        "  – VLAN":           sum(1 for i in interfaces if i['Type'] == 'vlan'),
        "  – Tunnel (VPN)":   sum(1 for i in interfaces if i['Type'] == 'tunnel'),
        "  – Aggregate":      sum(1 for i in interfaces if i['Type'] == 'aggregate'),
        "Zones":              len(zones),
        "VPN Phase1 Tunnels": len(phase1),
        "VPN Phase2 SAs":     len(phase2),
        "Static Routes":      len(routes),
        "SD-WAN Members":     len(sdwan['members']),
        "SD-WAN Health Checks": len(sdwan['health_checks']),
        "SD-WAN Services":    len(sdwan['services']),
        "Service Objects":     len(services),
        "  – With ports":     sum(1 for s in services if s['TCP Ports'] or s['UDP Ports']),
        "Service Groups":      len(svc_groups),
        "Firewall Policies":  len(policies),
        "  – Enabled":        sum(1 for p in policies if p['Status'] == 'enable'),
        "  – Disabled":       sum(1 for p in policies if p['Status'] == 'disable'),
        "  – Deny action":    sum(1 for p in policies if p['Action'] == 'deny'),
        "Address Objects":    len(addresses),
        "Address Groups":     len(addrgrps),
        "Admin Users":        len(admin_users),
    }

    for label, count in counts.items():
        print(f"  {label:<30} {count}")

    # ── Build workbook ────────────────────────────────────────────────────────
    print("\nBuilding workbook …")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    build_summary(wb, text, counts, str(input_path.name), pwd_policy)
    build_admin_users(wb, admin_users)
    build_interfaces(wb, interfaces)
    build_zones(wb, zones)
    build_vpn(wb, phase1, phase2)
    build_static_routes(wb, routes)
    build_sdwan(wb, sdwan)
    build_services(wb, services, svc_groups)
    build_policies(wb, policies)
    build_addresses(wb, addresses)
    build_addrgrps(wb, addrgrps)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
