#!/usr/bin/env python3
"""
xlsx_to_yaml.py
---------------
Generates conf.yaml from SNMP_Interface_Inventory.xlsx.

Each worksheet = one Fortigate/switch device.
Each data row  = one interface_config entry.

Column logic (flexible — add columns to the spreadsheet freely):
  - "Match Field", "Match Value"       → match_field / match_value YAML keys
  - "In Speed (bps)", "Out Speed (bps)"→ in_speed / out_speed YAML keys
  - "In Speed (Mbps)", "Out Speed (Mbps)" → skipped (derived columns)
  - Every other column                 → becomes a tag  (key = PascalCase column name)
                                          Empty cells are skipped (no blank tags)

Usage:
    python xlsx_to_yaml.py                         # uses default paths below
    python xlsx_to_yaml.py my.xlsx out.yaml        # custom paths
"""

import sys
import re
import textwrap
from pathlib import Path

import openpyxl

# ── Default paths ──────────────────────────────────────────────────────────────
DEFAULT_XLSX = Path(__file__).parent / "SNMP_Interface_Inventory.xlsx"
DEFAULT_YAML = Path(__file__).parent / "conf.yaml"

# ── Column roles ───────────────────────────────────────────────────────────────
# Maps spreadsheet header → YAML key (these never become tags)
CORE_COLUMNS = {
    "Match Field":     "match_field",
    "Match Value":     "match_value",
    "In Speed (bps)":  "in_speed",
    "Out Speed (bps)": "out_speed",
}

# Derived/formula columns — silently ignored
SKIP_COLUMNS = {"In Speed (Mbps)", "Out Speed (Mbps)"}

# ── Bare instances (no interface configs — polling-only devices) ───────────────
# These are not in the spreadsheet; update this list if devices are added/removed.
BARE_INSTANCES = [
    {"ip_address": "10.100.1.111", "community_string": "FCTGEAST-RO"},
    {"ip_address": "10.2.0.4",     "community_string": "FCTGAZURE-RO"},
    {"ip_address": "10.255.1.2",   "community_string": "FCTGDC-RO"},
    {"ip_address": "10.255.1.3",   "community_string": "FCTGDC-RO"},
    {"ip_address": "10.255.2.2",   "community_string": "FCTGDC-RO"},
    {"ip_address": "10.255.2.3",   "community_string": "FCTGDC-RO"},
    {"ip_address": "10.255.0.2",   "community_string": "FCTGHQ-RO"},
    {"ip_address": "10.255.0.3",   "community_string": "FCTGHQ-RO"},
    {"ip_address": "10.255.0.4",   "community_string": "FCTGHQ-RO"},
    {"ip_address": "10.255.0.5",   "community_string": "FCTGHQ-RO"},
    {"ip_address": "10.255.0.6",   "community_string": "FCTGHQ-RO"},
    {
        "ip_address": "10.255.0.7",
        "community_string": "FCTGHQ-RO",
        "loader": "core",
        "use_device_id_as_hostname": True,
        "profile": "fortinet-fortigate",
    },
]

# ── Helpers ────────────────────────────────────────────────────────────────────

def to_tag_key(col_name: str) -> str:
    """'Interface Zone' → 'InterfaceZone'  |  'VPN Destination' → 'VPNDestination'
    All-uppercase words (acronyms like VPN) are kept as-is."""
    return "".join(w if w.isupper() else w.capitalize() for w in col_name.split())


def parse_info_row(sheet) -> dict:
    """
    Row 1 contains label/value pairs side-by-side:
      col1=label, col2=value, col3=label, col4=value, ...
    Returns dict of {label: value}.
    """
    info = {}
    for col in range(1, sheet.max_column + 1, 2):
        label = sheet.cell(row=1, column=col).value
        value = sheet.cell(row=1, column=col + 1).value
        if label and value is not None:
            info[str(label).strip()] = str(value).strip()
    return info


def parse_headers(sheet) -> dict:
    """Row 2: returns {col_index: header_string} for non-empty cells."""
    return {
        col: str(sheet.cell(row=2, column=col).value).strip()
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(row=2, column=col).value is not None
    }


def clean_match_value(val) -> str:
    """Ensure match_value is always a plain string (Excel stores integers as floats)."""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def build_interface(row_vals: dict, headers: dict) -> dict | None:
    """
    Convert one spreadsheet row into an interface_config dict.
    Returns None if the row has no match_field (blank / summary row).
    """
    core = {}
    tags = []

    for col, col_name in headers.items():
        cell_val = row_vals.get(col)

        # ── Skip derived columns ──────────────────────────────────────
        if col_name in SKIP_COLUMNS:
            continue

        # ── Core YAML keys ────────────────────────────────────────────
        if col_name in CORE_COLUMNS:
            yaml_key = CORE_COLUMNS[col_name]
            if cell_val is not None:
                if yaml_key == "match_value":
                    cell_val = clean_match_value(cell_val)
                elif yaml_key in ("in_speed", "out_speed"):
                    cell_val = int(cell_val) if cell_val not in (None, "") else 0
                core[yaml_key] = cell_val

        # ── Tag columns (flexible — any new column becomes a tag) ─────
        else:
            if cell_val is not None and str(cell_val).strip():
                tags.append(f"{to_tag_key(col_name)}:{cell_val}")

    if not core.get("match_field"):
        return None

    return {
        "match_field": core.get("match_field"),
        "match_value": core.get("match_value"),
        "in_speed":    core.get("in_speed", 0),
        "out_speed":   core.get("out_speed", 0),
        "tags":        tags,
    }


def sheet_to_instance(sheet) -> dict:
    """Convert one worksheet into a Datadog SNMP instance dict."""
    info    = parse_info_row(sheet)
    headers = parse_headers(sheet)

    ip   = info.get("IP Address", "")
    snmp = info.get("SNMP Version", "v2c")
    auth = info.get("Auth / Community", "")

    # Base instance fields
    if snmp == "v3":
        instance = {
            "ip_address":              ip,
            "snmp_version":            3,
            "loader":                  "core",
            "use_device_id_as_hostname": True,
            "user":                    auth,
        }
    else:
        instance = {
            "ip_address":       ip,
            "community_string": auth,
        }

    # Interface configs
    iface_configs = []
    for row in range(3, sheet.max_row + 1):
        row_vals = {col: sheet.cell(row=row, column=col).value for col in headers}
        if all(v is None for v in row_vals.values()):
            continue
        iface = build_interface(row_vals, headers)
        if iface:
            iface_configs.append(iface)

    if iface_configs:
        instance["interface_configs"] = iface_configs

    return instance


# ── Custom YAML writer ─────────────────────────────────────────────────────────
# We write YAML manually to match Datadog's expected indentation style exactly.

def yaml_str(value: str) -> str:
    """Quote a string if it contains special characters, otherwise leave plain."""
    if re.search(r"[:{}\[\],&*?|<>=!%@`#]", value) or value in ("true", "false", "null", ""):
        return f"'{value}'"
    return value


def write_yaml(config: dict, path: Path):
    lines = []

    # ── init_config ────────────────────────────────────────────────────────────
    lines.append("init_config:")
    ic = config["init_config"]
    lines.append(f"    loader: {ic['loader']}")
    lines.append(f"    use_device_id_as_hostname: {str(ic['use_device_id_as_hostname']).lower()}")
    lines.append("    ping:")
    ping = ic["ping"]
    lines.append(f"        enabled: {str(ping['enabled']).lower()}")
    lines.append(f"        count: {ping['count']}")
    lines.append(f"        timeout: {ping['timeout']}")
    lines.append(f"        interval: {ping['interval']}")
    lines.append(f"    mibs_folder: '{ic['mibs_folder']}'")

    # ── instances ──────────────────────────────────────────────────────────────
    lines.append("instances:")

    for inst in config["instances"]:
        lines.append(f"  - ip_address: '{inst['ip_address']}'")

        # Auth
        if "community_string" in inst:
            lines.append(f"    community_string: '{inst['community_string']}'")
        if "snmp_version" in inst:
            lines.append(f"    snmp_version: {inst['snmp_version']}")
        if "loader" in inst:
            lines.append(f"    loader: {inst['loader']}")
        if "use_device_id_as_hostname" in inst:
            lines.append(f"    use_device_id_as_hostname: {str(inst['use_device_id_as_hostname']).lower()}")
        if "user" in inst:
            lines.append(f"    user: {inst['user']}")
        if "profile" in inst:
            lines.append(f"    profile: {inst['profile']}")

        # interface_configs
        if "interface_configs" in inst:
            lines.append("    interface_configs:")
            for iface in inst["interface_configs"]:
                mf  = iface["match_field"]
                mv  = iface["match_value"]
                # Index match values get quoted; name-based stay plain
                mv_str = f"'{mv}'" if mf == "index" else mv
                lines.append(f"    - match_field: {mf}")
                lines.append(f"      match_value: {mv_str}")
                lines.append(f"      in_speed: {iface['in_speed']}")
                lines.append(f"      out_speed: {iface['out_speed']}")
                if iface.get("tags"):
                    lines.append("      tags:")
                    for tag in iface["tags"]:
                        lines.append(f"      - {tag}")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Saved: {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    yaml_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_YAML

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    instances = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name}")
        instances.append(sheet_to_instance(ws))

    instances.extend(BARE_INSTANCES)

    config = {
        "init_config": {
            "loader": "core",
            "use_device_id_as_hostname": True,
            "ping": {"enabled": True, "count": 2, "timeout": 3000, "interval": 20},
            "mibs_folder": r"C:\ProgramData\Datadog\conf.d\snmp.d",
        },
        "instances": instances,
    }

    write_yaml(config, yaml_path)
    print("Done.")


if __name__ == "__main__":
    main()
