#!/usr/bin/env python3
"""
parse_full_config.py
---------------------
Reads a Fortigate full-configuration file ("show full-configuration" /
"get full-configuration") and creates a NEW spreadsheet pre-populated
with FriendlyName / VPNDestination values for every VPN tunnel found.

The output file is named:  <device>_<YYYYMMDD>.xlsx
and written to the same folder as the input config file (or --outdir).

KEY ADVANTAGE OVER parse_sys_interface.py
------------------------------------------
The full configuration file contains `set snmp-index` inside each
interface entry, so this script matches VPN tunnels to spreadsheet rows
by SNMP index directly — no positional guessing needed.

  Full config interface entry (abbreviated):
      edit "OPUS_to_SIFP"
          set type tunnel
          set ip 10.254.100.13 255.255.255.255
          set status up
          set snmp-index 59
      next

USAGE
-----
    python parse_full_config.py <config_file.conf> <device>
    python parse_full_config.py fctgfullconfig.txt 10.1.1.1
    python parse_full_config.py fctgfullconfig.txt "DC - 10.1.1.1" --dry-run

    <device>  is used as the sheet name and as part of the output filename.
              It can be a friendly name ("DC - 10.1.1.1") or just the IP.

OPTIONS
-------
    --xlsx PATH    Path to an existing spreadsheet to copy rows from
                   (default: SNMP_Interface_Inventory.xlsx in same folder).
                   If not found, a fresh spreadsheet is created instead.
    --outdir PATH  Directory for the output file (default: same folder as config file)
    --dry-run      Print planned output without writing any file
    --regen-yaml   After saving, run xlsx_to_yaml.py to regenerate conf.yaml

EXAMPLES
--------
    python parse_full_config.py fctgfullconfig.txt 10.1.1.1
    python parse_full_config.py fctgfullconfig.txt "DC - 10.1.1.1" --dry-run
    python parse_full_config.py fctgfullconfig.txt 10.1.1.1 --outdir ~/Desktop
    python parse_full_config.py fctgfullconfig.txt 10.1.1.1 --regen-yaml
"""

import sys
import re
import argparse
import subprocess
import ipaddress
from datetime import date
from pathlib import Path

import openpyxl

# ── Defaults ──────────────────────────────────────────────────────────────────
DEFAULT_XLSX    = Path(__file__).parent / "SNMP_Interface_Inventory.xlsx"
DEFAULT_YAML    = Path(__file__).parent / "conf.yaml"
XLSX_TO_YAML    = Path(__file__).parent / "xlsx_to_yaml.py"

# System tunnel names that are never IPsec VPN tunnels
SYSTEM_TUNNELS = {"naf.root", "l2t.root", "ssl.root"}

# ── OpCo subnet table (from CLAUDE.md) ────────────────────────────────────────
OPCO_SUBNETS = [
    ("10.0.0.0/16",  "FCTGHQ"),
    ("10.1.0.0/16",  "OPUS"),
    ("10.4.0.0/16",  "AIFP"),
    ("10.10.0.0/16", "BP"),
    ("10.14.0.0/16", "SMT"),
    ("10.18.0.0/16", "OI"),
    ("10.20.0.0/16", "TIFP"),
    ("10.24.0.0/16", "PFP"),
    ("10.28.0.0/16", "RIFP"),
    ("10.30.0.0/16", "SIFP"),
    ("10.75.0.0/16", "AR"),
]

# ── Output column layout ───────────────────────────────────────────────────────
# Columns xlsx_to_yaml.py treats as YAML core keys:
#   Match Field, Match Value, In Speed (bps), Out Speed (bps)
# All remaining columns become tags (PascalCase key).
OUTPUT_COLUMNS = [
    "Match Field",      # → match_field
    "Match Value",      # → match_value
    "In Speed (bps)",   # → in_speed
    "Out Speed (bps)",  # → out_speed
    "Interface Role",   # tag → InterfaceRole  (Primary / Secondary / Tertiary / VPN)
    "Provider",         # tag → Provider
    "OpCo",             # tag → OpCo
    "Friendly Name",    # tag → FriendlyName
    "VPN Destination",  # tag → VPNDestination
    "Interface Type",   # tag → InterfaceType
    "Interface Name",   # tag → InterfaceName
    "Interface IP",     # tag → InterfaceIp
    "Status",           # tag → Status
    "VDOM",             # tag → Vdom
]


# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_full_config_interfaces(text: str) -> list:
    """
    Parse `config system interface` block from a full FortiOS configuration.

    Returns an ordered list of dicts, one per interface:
        name        – interface edit name
        type        – interface type (physical, vlan, tunnel, aggregate, …)
        ip          – IP address string (e.g. "10.254.100.13"), "0.0.0.0" if none
        snmp_index  – int SNMP ifIndex (from `set snmp-index`), 0 if not present
        status      – "up" / "down" / "unknown"
        vdom        – vdom name
    """
    interfaces = []
    lines = text.splitlines()

    # Locate `config system interface` block using depth tracking
    start = next(
        (i for i, l in enumerate(lines) if l.strip() == 'config system interface'),
        None
    )
    if start is None:
        return interfaces

    depth = 1
    end = None
    for i in range(start + 1, len(lines)):
        s = lines[i].strip()
        if re.match(r'^config\b', s):
            depth += 1
        elif s == 'end':
            depth -= 1
            if depth == 0:
                end = i
                break

    if end is None:
        end = len(lines)

    # Parse individual interface entries (depth-1 edit/next pairs)
    current_name = None
    current_body = []

    def _commit(name, body_lines):
        body = '\n'.join(body_lines)
        def g(key):
            m = re.search(rf'set {key}\s+(\S+)', body)
            return m.group(1) if m else None
        ip_m = re.search(r'set ip\s+([\d.]+)\s+[\d.]+', body)
        snmp_m = re.search(r'set snmp-index\s+(\d+)', body)
        interfaces.append({
            'name':       name,
            'type':       g('type') or 'unknown',
            'ip':         ip_m.group(1) if ip_m else '0.0.0.0',
            'snmp_index': int(snmp_m.group(1)) if snmp_m else 0,
            'status':     g('status') or 'unknown',
            'vdom':       re.search(r'set vdom "([^"]+)"', body).group(1)
                          if re.search(r'set vdom "([^"]+)"', body) else 'root',
        })

    for line in lines[start + 1:end]:
        s = line.strip()
        if s.startswith('edit "') and current_name is None:
            m = re.search(r'edit "([^"]+)"', s)
            if m:
                current_name = m.group(1)
                current_body = []
        elif s == 'next' and current_name is not None:
            _commit(current_name, current_body)
            current_name = None
            current_body = []
        elif current_name is not None:
            current_body.append(line)

    # Commit last entry if file ended without 'next'
    if current_name:
        _commit(current_name, current_body)

    return interfaces


def filter_vpn_tunnels(interfaces: list) -> list:
    """
    Return only IPsec VPN tunnel interfaces, preserving parse order.

    Criteria:
      - type == 'tunnel'
      - name not in the known system tunnel set
      - ip is not 0.0.0.0 / None / empty  (system tunnels have no real IP)
    """
    return [
        iface for iface in interfaces
        if iface['type'] == 'tunnel'
        and iface['name'] not in SYSTEM_TUNNELS
        and iface['ip'] not in ('0.0.0.0', None, '')
    ]


# ── Device-level parsers ──────────────────────────────────────────────────────

def parse_snmp_community(text: str) -> str:
    """Return the first SNMP v2c community name defined in config system snmp community."""
    m = re.search(r'config system snmp community\n(.*?)\nend', text, re.DOTALL)
    if m:
        nm = re.search(r'set name "([^"]+)"', m.group(1))
        if nm:
            return nm.group(1)
    return ""


def parse_device_ip(text: str) -> str:
    """
    Best-effort: find the management IP from 'config system interface … edit "mgmt"'.
    Returns empty string if not found.
    """
    m = re.search(
        r'edit "mgmt"\s+(.*?)(?:\bnext\b)', text, re.DOTALL
    )
    if m:
        ip_m = re.search(r'set ip\s+([\d.]+)\s+[\d.]+', m.group(1))
        if ip_m:
            return ip_m.group(1)
    return ""


def derive_opco(ip_str: str) -> str:
    """Map a device or interface IP to an OpCo name using the CLAUDE.md subnet table."""
    if not ip_str:
        return ""
    try:
        addr = ipaddress.ip_address(ip_str)
        for cidr, name in OPCO_SUBNETS:
            if addr in ipaddress.ip_network(cidr):
                return name
    except ValueError:
        pass
    return ""


def parse_wan_interfaces(text: str) -> list:
    """
    Parse all physical/aggregate interfaces with role=wan (or monitor-bandwidth=enable)
    from config system interface.

    Returns a list of dicts with:
        name, type, ip, snmp_index, status, vdom,
        in_bps, out_bps,   ← estimated bandwidth converted to bps
        provider,           ← alias or description
    """
    wan_ifaces = []
    lines = text.splitlines()

    start = next(
        (i for i, l in enumerate(lines) if l.strip() == 'config system interface'), None
    )
    if start is None:
        return wan_ifaces

    depth = 1
    end = None
    for i in range(start + 1, len(lines)):
        s = lines[i].strip()
        if re.match(r'^config\b', s):
            depth += 1
        elif s == 'end':
            depth -= 1
            if depth == 0:
                end = i
                break
    if end is None:
        end = len(lines)

    current_name = None
    current_body = []

    def _commit(name, body_lines):
        body = '\n'.join(body_lines)
        def g(key):
            m = re.search(rf'set {key}\s+(\S+)', body)
            return m.group(1) if m else None
        def gq(key):
            m = re.search(rf'set {key}\s+"([^"]+)"', body)
            return m.group(1) if m else None

        role     = g('role') or ''
        mon_bw   = g('monitor-bandwidth') or 'disable'
        itype    = g('type') or 'unknown'

        # Include interfaces with role=wan OR monitor-bandwidth=enable
        if role != 'wan' and mon_bw != 'enable':
            return

        ip_m   = re.search(r'set ip\s+([\d.]+)\s+[\d.]+', body)
        snmp_m = re.search(r'set snmp-index\s+(\d+)', body)
        est_up = int(g('estimated-upstream-bandwidth')   or 0)
        est_dn = int(g('estimated-downstream-bandwidth') or 0)
        inbw   = int(g('inbandwidth')  or 0)
        outbw  = int(g('outbandwidth') or 0)

        # Prefer estimated-bandwidth; fall back to inbandwidth/outbandwidth (all in Kbps → bps)
        in_bps  = (est_dn  or inbw)  * 1000
        out_bps = (est_up  or outbw) * 1000

        # Provider: alias takes precedence, then description
        provider = (gq('alias') or gq('description') or '').strip()

        wan_ifaces.append({
            'name':       name,
            'type':       itype,
            'ip':         ip_m.group(1) if ip_m else '0.0.0.0',
            'snmp_index': int(snmp_m.group(1)) if snmp_m else 0,
            'status':     g('status') or 'unknown',
            'vdom':       gq('vdom') or 'root',
            'in_bps':     in_bps,
            'out_bps':    out_bps,
            'provider':   provider,
        })

    for line in lines[start + 1:end]:
        s = line.strip()
        if s.startswith('edit "') and current_name is None:
            m = re.search(r'edit "([^"]+)"', s)
            if m:
                current_name = m.group(1)
                current_body = []
        elif s == 'next' and current_name is not None:
            _commit(current_name, current_body)
            current_name = None
            current_body = []
        elif current_name is not None:
            current_body.append(line)

    if current_name:
        _commit(current_name, current_body)

    return wan_ifaces


def classify_wan_roles(wan_ifaces: list) -> list:
    """
    Assign Interface Role (Primary / Secondary / Tertiary / Other) to each WAN
    interface based on bandwidth, highest first.
    Mutates each dict in-place, adding a 'role' key.  Returns the list.
    """
    roles = ["Primary", "Secondary", "Tertiary"]
    # Sort by max of in/out bandwidth descending, then by snmp_index ascending for stable tie-break
    sorted_ifaces = sorted(wan_ifaces, key=lambda i: (-(max(i['in_bps'], i['out_bps'])), i['snmp_index']))
    for idx, iface in enumerate(sorted_ifaces):
        iface['role'] = roles[idx] if idx < len(roles) else "Other"
    return wan_ifaces


# ── Name helpers ──────────────────────────────────────────────────────────────

def derive_vpn_destination(tunnel_name: str) -> str:
    """
    Guess the VPNDestination tag value from the tunnel interface name.

    Strips trailing ordinal suffixes (_2, _2nd, _3rd …) then extracts
    the destination portion after the last _to_ or - separator.

    Examples:
        OPUS-FCTG        → FCTG
        OPUS-FCTG_2      → FCTG
        OPUS-Azure_2     → Azure
        OPUS_to_AIFP     → AIFP
        OPUS_to_AIFP_2   → AIFP
        OPUS_to_PFP_2nd  → PFP
        SMT_to_BIFP      → BIFP
        HQ_to_Azure_3rd  → Azure
    """
    cleaned = re.sub(r"_\d+(st|nd|rd|th)$", "", tunnel_name, flags=re.IGNORECASE)
    cleaned = re.sub(r"_\d+$", "", cleaned)

    m = re.search(r"_to_([A-Za-z0-9]+)$", cleaned, re.IGNORECASE)
    if m:
        return m.group(1)

    m = re.search(r"-([A-Za-z0-9]+)$", cleaned)
    if m:
        return m.group(1)

    return ""


# ── Spreadsheet helpers ────────────────────────────────────────────────────────

def find_sheet(workbook, device: str):
    """
    Locate the worksheet for <device>.
    <device> may be an exact sheet name or a substring (e.g. just the IP).
    Raises ValueError if not found.
    """
    device = device.strip()
    if device in workbook.sheetnames:
        return workbook[device], device
    for name in workbook.sheetnames:
        if device in name:
            return workbook[name], name
    raise ValueError(
        f"No sheet found for '{device}'.\n"
        f"Available sheets: {', '.join(workbook.sheetnames)}"
    )


def read_headers(sheet) -> dict:
    """
    Read row 2 as the column-header row.
    Returns dict: header_text → 1-based column index.
    """
    headers = {}
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=2, column=col).value
        if val:
            headers[str(val).strip()] = col
    return headers


def get_index_rows(sheet, headers) -> list:
    """
    Return all rows where match_field == 'index', sorted by index value (ascending).
    Each entry is (row_number, int_index).
    """
    mf_col = headers.get("Match Field")
    mv_col = headers.get("Match Value")
    if not mf_col or not mv_col:
        return []
    rows = []
    for row in range(3, sheet.max_row + 1):
        mf = sheet.cell(row=row, column=mf_col).value
        mv = sheet.cell(row=row, column=mv_col).value
        if str(mf).strip().lower() == "index" and mv is not None:
            try:
                rows.append((row, int(float(str(mv)))))
            except ValueError:
                pass
    return sorted(rows, key=lambda x: x[1])


def get_name_rows(sheet, headers) -> dict:
    """
    Return a dict of interface_name → row_number for match_field == 'name' rows.
    """
    mf_col = headers.get("Match Field")
    mv_col = headers.get("Match Value")
    if not mf_col or not mv_col:
        return {}
    result = {}
    for row in range(3, sheet.max_row + 1):
        mf = sheet.cell(row=row, column=mf_col).value
        mv = sheet.cell(row=row, column=mv_col).value
        if str(mf).strip().lower() == "name" and mv:
            result[str(mv).strip()] = row
    return result


# ── Output spreadsheet builder ────────────────────────────────────────────────

def build_output_sheet(wb, device: str, device_ip: str, community: str,
                       wan_ifaces: list, vpn_tunnels: list) -> None:
    """
    Create a sheet named <device> in wb, fully populated and compatible with
    xlsx_to_yaml.py.

    Row 1  – device info pairs read by parse_info_row():
                IP Address | <ip> | SNMP Version | v2c | Auth / Community | <community>
    Row 2  – column headers (OUTPUT_COLUMNS)
    Row 3+ – WAN interfaces (by role order), then VPN tunnels (by snmp_index)
    """
    if device in wb.sheetnames:
        del wb[device]

    ws = wb.create_sheet(device)

    # ── Row 1: device info ────────────────────────────────────────────────────
    row1 = [None] * (len(OUTPUT_COLUMNS))
    row1[0] = "IP Address";       row1[1] = device_ip
    row1[2] = "SNMP Version";     row1[3] = "v2c"
    row1[4] = "Auth / Community"; row1[5] = community
    ws.append(row1)

    # ── Row 2: column headers ─────────────────────────────────────────────────
    ws.append(OUTPUT_COLUMNS)

    # ── Helper: build a data row dict → ordered list ──────────────────────────
    col_idx = {c: i for i, c in enumerate(OUTPUT_COLUMNS)}

    def make_row(**kwargs):
        row = [None] * len(OUTPUT_COLUMNS)
        for k, v in kwargs.items():
            if k in col_idx:
                row[col_idx[k]] = v
        return row

    # ── WAN interfaces (sorted by role: Primary first) ────────────────────────
    role_order = {"Primary": 0, "Secondary": 1, "Tertiary": 2, "Other": 3}
    for iface in sorted(wan_ifaces, key=lambda i: (role_order.get(i.get('role', 'Other'), 9), i['snmp_index'])):
        ws.append(make_row(**{
            "Match Field":      "name",
            "Match Value":      iface['name'],
            "In Speed (bps)":   iface['in_bps'],
            "Out Speed (bps)":  iface['out_bps'],
            "Interface Role":   iface.get('role', ''),
            "Provider":         iface['provider'],
            "OpCo":             derive_opco(iface['ip']) or derive_opco(device_ip),
            "Interface Type":   iface['type'],
            "Interface Name":   iface['name'],
            "Interface IP":     iface['ip'],
            "Status":           iface['status'],
            "VDOM":             iface['vdom'],
        }))

    # ── VPN tunnels (sorted by snmp_index) ────────────────────────────────────
    for t in sorted(vpn_tunnels, key=lambda x: x['snmp_index']):
        dest = derive_vpn_destination(t['name'])
        ws.append(make_row(**{
            "Match Field":      "index",
            "Match Value":      t['snmp_index'],
            "In Speed (bps)":   0,
            "Out Speed (bps)":  0,
            "Interface Role":   "VPN",
            "OpCo":             derive_opco(device_ip),
            "Friendly Name":    t['name'],
            "VPN Destination":  dest,
            "Interface Type":   "tunnel",
            "Interface Name":   t['name'],
            "Interface IP":     t['ip'],
            "Status":           t['status'],
            "VDOM":             t['vdom'],
        }))

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [12, 12, 14, 14, 16, 20, 10, 30, 18, 14, 30, 15, 8, 8]
    for i, width in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width


def build_output_filename(device: str, outdir: Path) -> Path:
    """
    Construct the output path:  <outdir>/<device_safe>_<YYYYMMDD>.xlsx
    Strips characters that are illegal in filenames.
    """
    safe_device = re.sub(r'[\\/:*?"<>|]', '_', device).strip()
    datestamp   = date.today().strftime("%Y%m%d")
    return outdir / f"{safe_device}_{datestamp}.xlsx"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Parse a Fortigate full-config and create a new dated SNMP interface spreadsheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "config_file",
        help="Path to the full FortiOS configuration file (.conf / .txt)",
    )
    parser.add_argument(
        "device",
        help="Device label used as the sheet name and in the output filename "
             "(e.g. '10.1.1.1' or 'DC - 10.1.1.1')",
    )
    parser.add_argument(
        "--xlsx",
        default=str(DEFAULT_XLSX),
        metavar="PATH",
        help=f"Existing spreadsheet to copy non-tunnel rows from "
             f"(default: {DEFAULT_XLSX.name}).  Optional — skipped if not found.",
    )
    parser.add_argument(
        "--outdir",
        default=None,
        metavar="PATH",
        help="Output directory (default: same folder as the config file)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be written without creating any file",
    )
    parser.add_argument(
        "--regen-yaml",
        action="store_true",
        help="After saving, run xlsx_to_yaml.py to regenerate conf.yaml",
    )
    args = parser.parse_args()

    # ── Read config file ──────────────────────────────────────────────────────
    config_path = Path(args.config_file)
    if not config_path.exists():
        print(f"ERROR: Config file not found: {config_path}")
        sys.exit(1)

    print(f"Reading: {config_path}")
    config_text = config_path.read_text(encoding="utf-8", errors="replace")

    # ── Parse interfaces ──────────────────────────────────────────────────────
    all_interfaces = parse_full_config_interfaces(config_text)
    vpn_tunnels    = filter_vpn_tunnels(all_interfaces)

    print(f"Parsed {len(all_interfaces)} interfaces total.")
    print(f"Found {len(vpn_tunnels)} VPN tunnel(s) (type=tunnel, non-zero IP, non-system):\n")
    for t in vpn_tunnels:
        dest = derive_vpn_destination(t['name'])
        print(
            f"  snmp={t['snmp_index']:>3}  {t['name']:<40}  "
            f"ip={t['ip']:<18}  status={t['status']:<8}  "
            f"→ VPNDest: {dest or '(unrecognised)'}"
        )

    # ── Determine output path ─────────────────────────────────────────────────
    outdir     = Path(args.outdir) if args.outdir else config_path.parent
    out_path   = build_output_filename(args.device, outdir)

    print(f"\nOutput file will be: {out_path}")

    if args.dry_run:
        print("\n[DRY RUN — no file will be written]")
        return

    # ── Load source spreadsheet for non-tunnel rows (optional) ───────────────
    source_wb  = None
    xlsx_path  = Path(args.xlsx)
    if xlsx_path.exists():
        try:
            source_wb = openpyxl.load_workbook(xlsx_path)
            print(f"Source spreadsheet loaded: {xlsx_path.name}")
        except Exception as e:
            print(f"  WARNING: Could not open source spreadsheet ({e}) — skipping row copy.")
    else:
        print(f"  Source spreadsheet not found ({xlsx_path.name}) — creating fresh sheet.")

    # ── Build output workbook ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    print(f"\nBuilding sheet '{args.device}' ...")
    build_output_sheet(wb, args.device, all_interfaces, vpn_tunnels, source_wb)

    # ── Save ──────────────────────────────────────────────────────────────────
    outdir.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\nSaved: {out_path}")

    # ── Optionally regenerate conf.yaml ───────────────────────────────────────
    if args.regen_yaml:
        if XLSX_TO_YAML.exists():
            print(f"Regenerating conf.yaml via {XLSX_TO_YAML.name} ...")
            result = subprocess.run(
                [sys.executable, str(XLSX_TO_YAML), str(out_path), str(DEFAULT_YAML)],
                capture_output=True, text=True
            )
            if result.returncode == 0:
                print(result.stdout.strip())
                print(f"conf.yaml updated: {DEFAULT_YAML}")
            else:
                print(f"ERROR running xlsx_to_yaml.py:\n{result.stderr}")
        else:
            print(f"WARNING: xlsx_to_yaml.py not found at {XLSX_TO_YAML}")


if __name__ == "__main__":
    main()
